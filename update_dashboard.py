#!/usr/bin/env python3
"""
Robot de actualización del Tablero Ejecutivo — Agricola Guapa
----------------------------------------------------------------
Se ejecuta dentro de GitHub Actions. Hace 4 cosas:
  1. Se autentica ante Microsoft Graph con las credenciales de la app
     (usuario técnico, sin persona humana conectada).
  2. Busca y descarga el Excel más reciente desde el OneDrive indicado.
  3. Extrae y transforma los datos (misma lógica que usamos en el chat).
  4. Rellena la plantilla del tablero (dashboard_template.html) con los
     datos frescos y escribe index.html — el archivo que GitHub Pages sirve.

Variables de entorno esperadas (se configuran como "Secrets" en GitHub,
nunca quedan escritas en este archivo):
  AZURE_TENANT_ID       - Id. de directorio (inquilino)
  AZURE_CLIENT_ID       - Id. de aplicación (cliente)
  AZURE_CLIENT_SECRET   - El "Valor" del secreto de cliente
  ONEDRIVE_USER_EMAIL   - Correo del OneDrive donde vive el Excel
  EXCEL_FILE_NAME       - Nombre (o parte del nombre) del archivo a buscar
"""
import os
import sys
import json
import datetime
import requests
import openpyxl
from io import BytesIO

TENANT_ID = os.environ["AZURE_TENANT_ID"]
CLIENT_ID = os.environ["AZURE_CLIENT_ID"]
CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]
USER_EMAIL = os.environ["ONEDRIVE_USER_EMAIL"]
FILE_NAME = os.environ.get("EXCEL_FILE_NAME", "Indicadores_de_empacadora")

GRAPH = "https://graph.microsoft.com/v1.0"


def log(msg):
    print(f"[robot] {msg}", flush=True)


def get_access_token():
    log("Solicitando token de acceso a Microsoft...")
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }
    r = requests.post(url, data=data, timeout=30)
    if r.status_code != 200:
        log(f"ERROR al obtener token: {r.status_code} {r.text}")
        sys.exit(1)
    log("Token obtenido correctamente.")
    return r.json()["access_token"]


def find_excel_file(token):
    log(f"Buscando '{FILE_NAME}' en el OneDrive de {USER_EMAIL}...")
    headers = {"Authorization": f"Bearer {token}"}
    url = f"{GRAPH}/users/{USER_EMAIL}/drive/root/search(q='{FILE_NAME}')"
    r = requests.get(url, headers=headers, timeout=30)
    if r.status_code != 200:
        log(f"ERROR al buscar el archivo: {r.status_code} {r.text}")
        sys.exit(1)
    items = r.json().get("value", [])
    items = [i for i in items if i.get("name", "").lower().endswith((".xlsx", ".xlsm"))]
    if not items:
        log("ERROR: no encontré ningún archivo .xlsx/.xlsm que coincida.")
        sys.exit(1)
    # si hay varias coincidencias, toma la modificada más recientemente
    items.sort(key=lambda i: i.get("lastModifiedDateTime", ""), reverse=True)
    chosen = items[0]
    log(f"Encontrado: '{chosen['name']}' (modificado {chosen.get('lastModifiedDateTime')})")
    return chosen["id"]


def download_file(token, item_id):
    log("Descargando el archivo...")
    headers = {"Authorization": f"Bearer {token}"}
    url = f"{GRAPH}/users/{USER_EMAIL}/drive/items/{item_id}/content"
    r = requests.get(url, headers=headers, timeout=60)
    if r.status_code != 200:
        log(f"ERROR al descargar: {r.status_code} {r.text}")
        sys.exit(1)
    log(f"Descargado ({len(r.content)} bytes).")
    return BytesIO(r.content)


# ================= misma lógica de extracción que ya usamos en el chat =================

def is_err(v):
    return isinstance(v, str) and v.startswith("#")


def num(v, nd=3):
    if is_err(v) or v is None:
        return 0
    if isinstance(v, (int, float)):
        return round(v, nd)
    return v


def time_to_hours(t):
    if t is None:
        return 0
    if isinstance(t, datetime.time):
        return round(t.hour + t.minute / 60 + t.second / 3600, 3)
    if isinstance(t, (int, float)):
        return round(t, 3)
    return 0


def extract_data(file_bytes):
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    START_DATE = datetime.datetime(2023, 1, 2)

    ws0 = wb["Indicadores de planta"]
    last_real_date = None
    for r in ws0.iter_rows(min_row=3, values_only=True):
        fecha = r[0]
        if not isinstance(fecha, datetime.datetime):
            continue
        kg = r[4]
        if isinstance(kg, (int, float)) and kg > 0:
            if last_real_date is None or fecha > last_real_date:
                last_real_date = fecha
    END_DATE = last_real_date or datetime.datetime.now()
    log(f"Última fecha con producción real detectada: {END_DATE.date()}")

    # 1) Indicadores de planta (diario)
    ws = wb["Indicadores de planta"]
    fields1 = ["fecha", "semana", "mes", "anio", "kg_dia", "bines", "peso_bin", "bin_hora",
               "kg_hora", "pinas", "ocupacion", "rendimiento_planta", "descarte",
               "eficiencia_planta", "eficiencia_mo", "rendimiento_mo", "operarios",
               "kg_hombre", "uni_hombre"]
    rows1 = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        fecha = r[0]
        if not isinstance(fecha, datetime.datetime):
            continue
        if fecha < START_DATE or fecha > END_DATE:
            continue
        kg = r[4]
        if is_err(kg):
            kg = 0
        rows1.append([
            fecha.strftime("%Y-%m-%d"), int(r[1]) if r[1] else 0, r[2], int(r[3]) if r[3] else 0,
            num(kg, 0), num(r[5], 0), num(r[6], 1), num(r[7], 2), num(r[8], 1), num(r[9], 0),
            num(r[12], 3), num(r[13], 4), num(r[14], 4), num(r[15], 4), num(r[16], 4),
            num(r[17], 4), num(r[18], 0), num(r[19], 1), num(r[20], 1),
        ])
    log(f"Indicadores de planta: {len(rows1)} filas")

    # 2) Horas extras (todos los años disponibles)
    fields2 = ["anio", "semana", "horas_extras_total", "kilos_procesados", "recargo_nocturno",
               "horas_extra_ordinarias", "recargo_festivo", "horas_extra_festivas"]
    rows2 = []
    for name in wb.sheetnames:
        if not name.strip().lower().startswith("horas extras"):
            continue
        import re
        m = re.search(r"(\d{4})", name)
        if not m:
            continue
        year = int(m.group(1))
        ws = wb[name]
        header = [c.value if hasattr(c, "value") else c for c in next(ws.iter_rows(min_row=1, max_row=1))]
        has_full_cols = len(header) > 6
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[1] is None:
                continue
            if has_full_cols and len(r) > 7:
                rows2.append([year, int(r[1]), num(r[2], 1), num(r[3], 0), num(r[4], 1),
                              num(r[5], 1), num(r[6], 1), num(r[7], 1)])
            else:
                rows2.append([year, int(r[1]), num(r[2], 1), num(r[3], 0), 0,
                              num(r[4], 1), 0, num(r[5], 1)])
    log(f"Horas extras: {len(rows2)} filas")

    # 3) Tiempomuertoanual — detección robusta de columnas (soporta causas nuevas)
    ws = wb["Tiempomuertoanual"]
    headers = [c for c in ws.iter_rows(min_row=1, max_row=1, values_only=True)][0]
    tm_total_idx = None
    horas_trab_idx = None
    for idx, hd in enumerate(headers):
        hdn = str(hd).lower() if hd else ""
        if "tiempo muerto" in hdn and tm_total_idx is None:
            tm_total_idx = idx
        if "horas trabajadas" in hdn:
            horas_trab_idx = idx
    causa_names = [headers[i].strip() for i in range(2, tm_total_idx)]
    fields3 = ["fecha", "semana"] + causa_names + ["tiempo_muerto_total_hr", "horas_trabajadas"]
    rows3 = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        fecha = r[0]
        if not isinstance(fecha, datetime.datetime):
            continue
        if fecha < START_DATE or fecha > END_DATE:
            continue
        causas_vals = [time_to_hours(r[i]) if i < len(r) else 0 for i in range(2, tm_total_idx)]
        tm_total = time_to_hours(r[tm_total_idx]) if tm_total_idx < len(r) else 0
        ht = time_to_hours(r[horas_trab_idx]) if horas_trab_idx and horas_trab_idx < len(r) else 0
        rows3.append([fecha.strftime("%Y-%m-%d"), int(r[1]) if r[1] else 0] + causas_vals + [tm_total, ht])
    log(f"Tiempomuertoanual: {len(rows3)} filas, {len(causa_names)} causas ({causa_names})")

    # 4) Consolidado Post-cosecha
    ws = wb["Consolidado Post-cosecha"]
    headers = [c for c in ws.iter_rows(min_row=1, max_row=1, values_only=True)][0]
    causa_names_pc = [headers[i].strip() for i in range(1, 12)]
    fields4 = ["fecha"] + causa_names_pc + ["total_merma", "total_cosecha", "total_despacho"]
    rows4 = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        fecha = r[0]
        if not isinstance(fecha, datetime.datetime):
            continue
        if fecha < START_DATE or fecha > END_DATE:
            continue
        causas_vals = [num(r[i], 0) if i < len(r) else 0 for i in range(1, 12)]
        rows4.append([fecha.strftime("%Y-%m-%d")] + causas_vals + [num(r[12], 0), num(r[13], 0), num(r[14], 0)])
    log(f"Consolidado Post-cosecha: {len(rows4)} filas")

    # 5) TiempoE (fuente exacta de "Datos Eduar")
    ws = wb["TiempoE"]
    rowsE = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        semana, anio, kg, horas_totales, tiempo_muerto, horas_netas = r[0], r[1], r[2], r[3], r[4], r[5]
        if not isinstance(semana, (int, float)) or not isinstance(anio, (int, float)):
            continue
        if anio not in (2023, 2024, 2025, 2026, 2027):
            continue
        rowsE.append([int(anio), int(semana), kg or 0, horas_totales or 0, tiempo_muerto or 0, horas_netas or 0])
    while rowsE and rowsE[-1][5] == 0:
        rowsE.pop()
    log(f"TiempoE: {len(rowsE)} filas")

    return {
        "daily": {"fields": fields1, "rows": rows1},
        "horas_extras": {"fields": fields2, "rows": rows2},
        "tiempo_muerto": {"fields": fields3, "rows": rows3},
        "post_cosecha": {"fields": fields4, "rows": rows4},
        "tiempo_e": {"fields": ["anio", "semana", "kg", "horas_totales", "tiempo_muerto", "horas_netas"], "rows": rowsE},
    }


def main():
    token = get_access_token()
    item_id = find_excel_file(token)
    file_bytes = download_file(token, item_id)

    log("Extrayendo datos del Excel...")
    data = extract_data(file_bytes)

    log("Rellenando la plantilla del tablero...")
    with open("dashboard_template.html", "r", encoding="utf-8") as f:
        html = f.read()
    html = html.replace("__DATA_PLACEHOLDER__", json.dumps(data, ensure_ascii=False, separators=(",", ":")))

    with open("index.html", "w", encoding="utf-8") as f:
        f.write(html)
    log(f"index.html generado correctamente ({len(html)} caracteres).")


if __name__ == "__main__":
    main()
